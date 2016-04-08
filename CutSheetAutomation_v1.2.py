#script to read linear hardware inventory and compile provisioning/cut sheets for data center migrations
from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string, cell
from openpyxl.styles import Font, Style, fills, PatternFill, Color, Alignment, Border, Side
import os, ctypes, getpass, easygui

#desired working directory
working_directory = os.path.join('c:\users\\', os.getenv('username'), 'Desktop\ProvisioningSheets')

#check to see if directory exists and create it if not
if not os.path.exists(working_directory):
    os.makedirs(working_directory)
	
#Change the working directory
os.chdir(working_directory)

#Pop up to move the linear inventory to the working directory found in working_directory

ctypes.windll.user32.MessageBoxA(0, 'Move linear inventory to ' + working_directory + ' then select ok', 'Move Linear Inventory', 1)

#get linear inventory name and load as a workbook

while True:
	ctypes.windll.user32.MessageBoxA(0, 'Open your Linear Inventory', 'Open', 1)
	linear = easygui.fileopenbox()
	try:
		linear_wb = load_workbook(filename = linear, use_iterators = True)
		break
	except:
		ctypes.windll.user32.MessageBoxA(0, 'Please enter a valid linear inventory name', 'Import Error!', 1)
		continue
active_sheet = linear_wb.get_sheet_names()[0]
working_sheet = linear_wb.get_sheet_by_name(active_sheet) 

#create the provisioning sheets compile workbook
cust_name = raw_input('Enter the customer name with no punctuation or spaces (eg. Childrens, MacArthur, SunCountry etc.)\n')
provision_name = cust_name+'_ProvisionCutSheet.xlsx'
provision_wb = Workbook()
provision_sheet = provision_wb.active

#define the increase constant which is equal to the number of rows that can fit on a single sheet
i_constant = 68

#Cell Styling
boldfont = Font(bold = True)
#Orange
cellstyle1 = Style(font = boldfont, fill = PatternFill(patternType = 'solid', fgColor = Color('FF9900')))
#Blue
cellstyle2 = Style(font = boldfont, fill = PatternFill(patternType = 'solid', fgColor = Color('3366FF')))
#Green
cellstyle3 = Style(font = boldfont, fill = PatternFill(patternType = 'solid', fgColor = Color('99CC00')))

#Prepare the cutsheet with standard header values
#create the cell indexes for cut sheet dynamic values
cutnum = list()

for i in range (1, 69):
	cutnum.append(i)
	
#Hard code other header values and bold the cell contents

for i in range (0, working_sheet.get_highest_row() - 1):
	#column A
	provision_sheet['A' + str(cutnum[0])] = 'Device Name'
	provision_sheet['A' + str(cutnum[0])].style = cellstyle1
	provision_sheet['A' + str(cutnum[1])] = 'Asset Tag#'
	provision_sheet['A' + str(cutnum[1])].style = cellstyle1
	provision_sheet['A' + str(cutnum[2])] = 'Serial Number'
	provision_sheet['A' + str(cutnum[2])].style = cellstyle1
	provision_sheet['A' + str(cutnum[3])] = 'Manufacturer'
	provision_sheet['A' + str(cutnum[3])].style = cellstyle1
	provision_sheet['A' + str(cutnum[4])] = 'Device Type'
	provision_sheet['A' + str(cutnum[4])].style = cellstyle1
	provision_sheet['A' + str(cutnum[5])] = 'System Model#'
	provision_sheet['A' + str(cutnum[5])].style = cellstyle1
	provision_sheet['A' + str(cutnum[6])] = 'Total RMU'
	provision_sheet['A' + str(cutnum[6])].style = cellstyle1
	provision_sheet['A' + str(cutnum[7])] = 'Has iLO Port'
	provision_sheet['A' + str(cutnum[7])].style = cellstyle1
	provision_sheet['A' + str(cutnum[8])] = 'Number of Analog Connections'
	provision_sheet['A' + str(cutnum[8])].style = cellstyle1
	provision_sheet['A' + str(cutnum[10])] = 'Source'
	provision_sheet['A' + str(cutnum[10])].style = cellstyle2
	provision_sheet['A' + str(cutnum[12])] = 'Source Cabinet'
	provision_sheet['A' + str(cutnum[12])].style = cellstyle2
	provision_sheet['A' + str(cutnum[13])] = 'Source Start RU'
	provision_sheet['A' + str(cutnum[13])].style = cellstyle2
	provision_sheet['A' + str(cutnum[14])] = 'Source End RU'
	provision_sheet['A' + str(cutnum[14])].style = cellstyle2
	provision_sheet['A' + str(cutnum[15])] = 'Source Mount Position'
	provision_sheet['A' + str(cutnum[15])].style = cellstyle2
	provision_sheet['A' + str(cutnum[17])] = 'Source PS Quantity'
	provision_sheet['A' + str(cutnum[17])].style = cellstyle2
	provision_sheet['A' + str(cutnum[18])] = 'Source PS1'
	provision_sheet['A' + str(cutnum[18])].style = cellstyle2
	provision_sheet['A' + str(cutnum[19])] = 'Source PS2'
	provision_sheet['A' + str(cutnum[19])].style = cellstyle2
	provision_sheet['A' + str(cutnum[20])] = 'Source PS3'
	provision_sheet['A' + str(cutnum[20])].style = cellstyle2
	provision_sheet['A' + str(cutnum[21])] = 'Source PS4'
	provision_sheet['A' + str(cutnum[21])].style = cellstyle2
	provision_sheet['A' + str(cutnum[23])] = 'Source IP 1'
	provision_sheet['A' + str(cutnum[23])].style = cellstyle2
	provision_sheet['A' + str(cutnum[24])] = 'Source IP 2'
	provision_sheet['A' + str(cutnum[24])].style = cellstyle2
	provision_sheet['A' + str(cutnum[26])] = 'Source Device Port'
	provision_sheet['A' + str(cutnum[26])].style = cellstyle2
	provision_sheet['A' + str(cutnum[52])] = 'Critical Comments/Notes'
	provision_sheet['A' + str(cutnum[52])].style = cellstyle1
	provision_sheet['A' + str(cutnum[63])] = 'Data Collector/Validator'
	provision_sheet['A' + str(cutnum[63])].style = cellstyle1
	provision_sheet['A' + str(cutnum[64])] = 'Deinstaller/Mover'
	provision_sheet['A' + str(cutnum[64])].style = cellstyle1
	provision_sheet['A' + str(cutnum[65])] = 'Installer'
	provision_sheet['A' + str(cutnum[65])].style = cellstyle1
	provision_sheet['A' + str(cutnum[66])] = 'Completed By'
	provision_sheet['A' + str(cutnum[66])].style = cellstyle1
	provision_sheet['A' + str(cutnum[67])] = 'Reviewer'
	provision_sheet['A' + str(cutnum[67])].style = cellstyle1
	
	#Column B
	provision_sheet['B' + str(cutnum[26])] = 'Source Termination Port'
	provision_sheet['B' + str(cutnum[26])].style = cellstyle2
	#Column C
	provision_sheet['C' + str(cutnum[26])] = 'Cable Media'
	provision_sheet['C' + str(cutnum[26])].style = cellstyle1
	#Column D
	provision_sheet['D' + str(cutnum[0])] = 'Previous Device Name'
	provision_sheet['D' + str(cutnum[0])].style = cellstyle1
	provision_sheet['D' + str(cutnum[1])] = 'Destination'
	provision_sheet['D' + str(cutnum[1])].style = cellstyle1
	provision_sheet['D' + str(cutnum[2])] = 'Move Date'
	provision_sheet['D' + str(cutnum[2])].style = cellstyle1
	provision_sheet['D' + str(cutnum[3])] = 'Event#'
	provision_sheet['D' + str(cutnum[3])].style = cellstyle1
	provision_sheet['D' + str(cutnum[4])] = 'Truck #'
	provision_sheet['D' + str(cutnum[4])].style = cellstyle1
	provision_sheet['D' + str(cutnum[5])] = 'Move Team'
	provision_sheet['D' + str(cutnum[5])].style = cellstyle1
	provision_sheet['D' + str(cutnum[6])] = 'Owner/Support'
	provision_sheet['D' + str(cutnum[6])].style = cellstyle1
	provision_sheet['D' + str(cutnum[7])] = 'Phone Number'
	provision_sheet['D' + str(cutnum[7])].style = cellstyle1
	provision_sheet['D' + str(cutnum[10])] = 'Destination'
	provision_sheet['D' + str(cutnum[10])].style = cellstyle3
	provision_sheet['D' + str(cutnum[12])] = 'Destination Cabinet'
	provision_sheet['D' + str(cutnum[12])].style = cellstyle3
	provision_sheet['D' + str(cutnum[13])] = 'Destination Start RU'
	provision_sheet['D' + str(cutnum[13])].style = cellstyle3
	provision_sheet['D' + str(cutnum[14])] = 'Destination End RU'
	provision_sheet['D' + str(cutnum[14])].style = cellstyle3
	provision_sheet['D' + str(cutnum[15])] = 'Destination Mount Position'
	provision_sheet['D' + str(cutnum[15])].style = cellstyle3
	provision_sheet['D' + str(cutnum[17])] = 'Destination PS Quantity'
	provision_sheet['D' + str(cutnum[17])].style = cellstyle3
	provision_sheet['D' + str(cutnum[18])] = 'Destination PS1'
	provision_sheet['D' + str(cutnum[18])].style = cellstyle3
	provision_sheet['D' + str(cutnum[19])] = 'Destination PS2'
	provision_sheet['D' + str(cutnum[19])].style = cellstyle3
	provision_sheet['D' + str(cutnum[20])] = 'Destination PS3'
	provision_sheet['D' + str(cutnum[20])].style = cellstyle3
	provision_sheet['D' + str(cutnum[21])] = 'Destination PS4'
	provision_sheet['D' + str(cutnum[21])].style = cellstyle3
	provision_sheet['D' + str(cutnum[23])] = 'Destination IP 1'
	provision_sheet['D' + str(cutnum[23])].style = cellstyle3
	provision_sheet['D' + str(cutnum[24])] = 'Destination IP 2'
	provision_sheet['D' + str(cutnum[24])].style = cellstyle3
	provision_sheet['D' + str(cutnum[26])] = 'Destination Device Port'
	provision_sheet['D' + str(cutnum[26])].style = cellstyle3
	provision_sheet['D' + str(cutnum[67])] = 'Date'
	provision_sheet['D' + str(cutnum[67])].style = cellstyle1
	provision_sheet['D' + str(cutnum[63])] = 'Date'
	provision_sheet['D' + str(cutnum[63])].style = cellstyle1
	provision_sheet['D' + str(cutnum[64])] = 'Date'
	provision_sheet['D' + str(cutnum[64])].style = cellstyle1
	provision_sheet['D' + str(cutnum[65])] = 'Date'
	provision_sheet['D' + str(cutnum[65])].style = cellstyle1
	provision_sheet['D' + str(cutnum[66])] = 'Date'
	provision_sheet['D' + str(cutnum[66])].style = cellstyle1
	
	#Column E
	provision_sheet['E' + str(cutnum[26])] = 'Destination Termination Port'
	provision_sheet['E' + str(cutnum[26])].style = cellstyle3
	
	#Merge Cells for Critical Comments Section
	provision_sheet.merge_cells('A' + str(cutnum[53]) + ':' + 'E' + str(cutnum[56]))
	
	#increment index values by 68
	cutnum[:] = [i + i_constant for i in cutnum]
	
#create dictionary keys based on column headers
keys = list()
h_col = working_sheet.get_highest_column()
for row in working_sheet['A1': get_column_letter(h_col)+'1']:
	for cell in row:
		keys.append(cell.value)

#Cell indexes for cut sheet dynamic values
cutnum2 = list()

for i in range (1, 69):
	cutnum2.append(i)
		
#Loop to create temporary dictionaries and write to the cut sheet
index = 0
value_dict = {}
for row in working_sheet['A2': get_column_letter(h_col) + str(working_sheet.get_highest_row())]:
	print 'Working...'
	for cell in row:
		value_dict.update({keys[index]:cell.value})
		if index < h_col - 1:
			index = index + 1
		else:
			index = 0
	
	# write dictionary values to the cutsheet excel document
	#provision_sheet
	
	#Column B Non Port Information
	provision_sheet['B' + str(cutnum2[0])] = value_dict.get('Device Name', None)
	provision_sheet['B' + str(cutnum2[1])] = value_dict.get('Asset Tag#', None)
	provision_sheet['B' + str(cutnum2[2])] = value_dict.get('Serial Number', None)
	provision_sheet['B' + str(cutnum2[3])] = value_dict.get('Manufacturer', None)
	provision_sheet['B' + str(cutnum2[4])] = value_dict.get('Device Type', None)
	provision_sheet['B' + str(cutnum2[5])] = value_dict.get('System Model#', None)
	provision_sheet['B' + str(cutnum2[6])] = value_dict.get('Total RMU', None)
	provision_sheet['B' + str(cutnum2[7])] = value_dict.get('Has iLO Port', None)
	provision_sheet['B' + str(cutnum2[8])] = value_dict.get('Num_Analog', None)
	provision_sheet['B' + str(cutnum2[12])] = value_dict.get('Source Cabinet', None)
	provision_sheet['B' + str(cutnum2[13])] = value_dict.get('Source Start RU', None)
	provision_sheet['B' + str(cutnum2[14])] = value_dict.get('Source End RU', None)
	provision_sheet['B' + str(cutnum2[15])] = value_dict.get('Source Mount Position', None)
	provision_sheet['B' + str(cutnum2[17])] = value_dict.get('Source PS Qty', None)
	provision_sheet['B' + str(cutnum2[18])] = value_dict.get('Source PS1', None)
	provision_sheet['B' + str(cutnum2[19])] = value_dict.get('Source PS2', None)
	provision_sheet['B' + str(cutnum2[20])] = value_dict.get('Source PS3', None)
	provision_sheet['B' + str(cutnum2[21])] = value_dict.get('Source PS4', None)
	provision_sheet['B' + str(cutnum2[23])] = value_dict.get('Source IP1', None)
	provision_sheet['B' + str(cutnum2[24])] = value_dict.get('Source IP2', None)
		
	#Column E Non Port Information
	provision_sheet['E' + str(cutnum2[0])] = value_dict.get('Previous Device Name', None)
	provision_sheet['E' + str(cutnum2[1])] = value_dict.get('Destination Location', None)
	provision_sheet['E' + str(cutnum2[2])] = value_dict.get('Move Date', None)
	provision_sheet['E' + str(cutnum2[3])] = value_dict.get('Event#', None)
	provision_sheet['E' + str(cutnum2[4])] = value_dict.get('Truck #', None)
	provision_sheet['E' + str(cutnum2[5])] = value_dict.get('Move Team', None)
	provision_sheet['E' + str(cutnum2[6])] = value_dict.get('Owner/Support', None)
	provision_sheet['E' + str(cutnum2[7])] = value_dict.get('Phone Number', None)
	provision_sheet['E' + str(cutnum2[12])] = value_dict.get('Destination Cabinet', None)
	provision_sheet['E' + str(cutnum2[13])] = value_dict.get('Destination Start RU', None)
	provision_sheet['E' + str(cutnum2[14])] = value_dict.get('Destination End RU', None)
	provision_sheet['E' + str(cutnum2[15])] = value_dict.get('Destination Mount Position', None)
	provision_sheet['E' + str(cutnum2[17])] = value_dict.get('Destination PS Qty', None)
	provision_sheet['E' + str(cutnum2[18])] = value_dict.get('Destination PS1', None)
	provision_sheet['E' + str(cutnum2[19])] = value_dict.get('Destination PS2', None)
	provision_sheet['E' + str(cutnum2[20])] = value_dict.get('Destination PS3', None)
	provision_sheet['E' + str(cutnum2[21])] = value_dict.get('Destination PS4', None)
	provision_sheet['E' + str(cutnum2[23])] = value_dict.get('Destination IP1', None)
	provision_sheet['E' + str(cutnum2[24])] = value_dict.get('Destination IP2', None)
		
	#Column A Non Port Information
	provision_sheet['A' + str(cutnum2[53])] = value_dict.get('Critical Comments/Notes', None)
	
	#Column A Port Information
	provision_sheet['A' + str(cutnum2[27])] = value_dict.get('Source Device Port 1', None)
	provision_sheet['A' + str(cutnum2[28])] = value_dict.get('Source Device Port 2', None)
	provision_sheet['A' + str(cutnum2[29])] = value_dict.get('Source Device Port 3', None)
	provision_sheet['A' + str(cutnum2[30])] = value_dict.get('Source Device Port 4', None)
	provision_sheet['A' + str(cutnum2[31])] = value_dict.get('Source Device Port 5', None)
	provision_sheet['A' + str(cutnum2[32])] = value_dict.get('Source Device Port 6', None)
	provision_sheet['A' + str(cutnum2[33])] = value_dict.get('Source Device Port 7', None)
	provision_sheet['A' + str(cutnum2[34])] = value_dict.get('Source Device Port 8', None)
	provision_sheet['A' + str(cutnum2[35])] = value_dict.get('Source Device Port 9', None)
	provision_sheet['A' + str(cutnum2[36])] = value_dict.get('Source Device Port 10', None)
	provision_sheet['A' + str(cutnum2[37])] = value_dict.get('Source Device Port 11', None)
	provision_sheet['A' + str(cutnum2[38])] = value_dict.get('Source Device Port 12', None)
	provision_sheet['A' + str(cutnum2[39])] = value_dict.get('Source Device Port 13', None)
	provision_sheet['A' + str(cutnum2[40])] = value_dict.get('Source Device Port 14', None)
	provision_sheet['A' + str(cutnum2[41])] = value_dict.get('Source Device Port 15', None)
	provision_sheet['A' + str(cutnum2[42])] = value_dict.get('Source Device Port 16', None)
	provision_sheet['A' + str(cutnum2[43])] = value_dict.get('Source Device Port 17', None)
	provision_sheet['A' + str(cutnum2[44])] = value_dict.get('Source Device Port 18', None)
	provision_sheet['A' + str(cutnum2[45])] = value_dict.get('Source Device Port 19', None)
	provision_sheet['A' + str(cutnum2[46])] = value_dict.get('Source Device Port 20', None)
	provision_sheet['A' + str(cutnum2[47])] = value_dict.get('Source Device Port 21', None)
	provision_sheet['A' + str(cutnum2[48])] = value_dict.get('Source Device Port 22', None)
	provision_sheet['A' + str(cutnum2[49])] = value_dict.get('Source Device Port 23', None)
	provision_sheet['A' + str(cutnum2[50])] = value_dict.get('Source Device Port 24', None)
	
	#Column B Port Information
	provision_sheet['B' + str(cutnum2[27])] = value_dict.get('Source Termination Port 1', None)
	provision_sheet['B' + str(cutnum2[27])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[28])] = value_dict.get('Source Termination Port 2', None)
	provision_sheet['B' + str(cutnum2[28])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[29])] = value_dict.get('Source Termination Port 3', None)
	provision_sheet['B' + str(cutnum2[29])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[30])] = value_dict.get('Source Termination Port 4', None)
	provision_sheet['B' + str(cutnum2[30])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[31])] = value_dict.get('Source Termination Port 5', None)
	provision_sheet['B' + str(cutnum2[31])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[32])] = value_dict.get('Source Termination Port 6', None)
	provision_sheet['B' + str(cutnum2[32])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[33])] = value_dict.get('Source Termination Port 7', None)
	provision_sheet['B' + str(cutnum2[33])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[34])] = value_dict.get('Source Termination Port 8', None)
	provision_sheet['B' + str(cutnum2[34])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[35])] = value_dict.get('Source Termination Port 9', None)
	provision_sheet['B' + str(cutnum2[35])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[36])] = value_dict.get('Source Termination Port 10', None)
	provision_sheet['B' + str(cutnum2[36])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[37])] = value_dict.get('Source Termination Port 11', None)
	provision_sheet['B' + str(cutnum2[37])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[38])] = value_dict.get('Source Termination Port 12', None)
	provision_sheet['B' + str(cutnum2[38])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[39])] = value_dict.get('Source Termination Port 13', None)
	provision_sheet['B' + str(cutnum2[39])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[40])] = value_dict.get('Source Termination Port 14', None)
	provision_sheet['B' + str(cutnum2[40])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[41])] = value_dict.get('Source Termination Port 15', None)
	provision_sheet['B' + str(cutnum2[41])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[42])] = value_dict.get('Source Termination Port 16', None)
	provision_sheet['B' + str(cutnum2[42])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[43])] = value_dict.get('Source Termination Port 17', None)
	provision_sheet['B' + str(cutnum2[43])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[44])] = value_dict.get('Source Termination Port 18', None)
	provision_sheet['B' + str(cutnum2[44])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[45])] = value_dict.get('Source Termination Port 19', None)
	provision_sheet['B' + str(cutnum2[45])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[46])] = value_dict.get('Source Termination Port 20', None)
	provision_sheet['B' + str(cutnum2[46])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[47])] = value_dict.get('Source Termination Port 21', None)
	provision_sheet['B' + str(cutnum2[47])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[48])] = value_dict.get('Source Termination Port 22', None)
	provision_sheet['B' + str(cutnum2[48])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[49])] = value_dict.get('Source Termination Port 23', None)
	provision_sheet['B' + str(cutnum2[49])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['B' + str(cutnum2[50])] = value_dict.get('Source Termination Port 24', None)
	provision_sheet['B' + str(cutnum2[50])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	
	#Column C Port Information
	provision_sheet['C' + str(cutnum2[27])] = value_dict.get('Cable Media 1', None)
	provision_sheet['C' + str(cutnum2[27])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[28])] = value_dict.get('Cable Media 2', None)
	provision_sheet['C' + str(cutnum2[28])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[29])] = value_dict.get('Cable Media 3', None)
	provision_sheet['C' + str(cutnum2[29])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[30])] = value_dict.get('Cable Media 4', None)
	provision_sheet['C' + str(cutnum2[30])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[31])] = value_dict.get('Cable Media 5', None)
	provision_sheet['C' + str(cutnum2[31])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[32])] = value_dict.get('Cable Media 6', None)
	provision_sheet['C' + str(cutnum2[32])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[33])] = value_dict.get('Cable Media 7', None)
	provision_sheet['C' + str(cutnum2[33])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[34])] = value_dict.get('Cable Media 8', None)
	provision_sheet['C' + str(cutnum2[34])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[35])] = value_dict.get('Cable Media 9', None)
	provision_sheet['C' + str(cutnum2[35])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[36])] = value_dict.get('Cable Media 10', None)
	provision_sheet['C' + str(cutnum2[36])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[37])] = value_dict.get('Cable Media 11', None)
	provision_sheet['C' + str(cutnum2[37])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[38])] = value_dict.get('Cable Media 12', None)
	provision_sheet['C' + str(cutnum2[38])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[39])] = value_dict.get('Cable Media 13', None)
	provision_sheet['C' + str(cutnum2[39])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[40])] = value_dict.get('Cable Media 14', None)
	provision_sheet['C' + str(cutnum2[40])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[41])] = value_dict.get('Cable Media 15', None)
	provision_sheet['C' + str(cutnum2[41])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[42])] = value_dict.get('Cable Media 16', None)
	provision_sheet['C' + str(cutnum2[42])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[43])] = value_dict.get('Cable Media 17', None)
	provision_sheet['C' + str(cutnum2[43])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[44])] = value_dict.get('Cable Media 18', None)
	provision_sheet['C' + str(cutnum2[44])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[45])] = value_dict.get('Cable Media 19', None)
	provision_sheet['C' + str(cutnum2[45])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[46])] = value_dict.get('Cable Media 20', None)
	provision_sheet['C' + str(cutnum2[46])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[47])] = value_dict.get('Cable Media 21', None)
	provision_sheet['C' + str(cutnum2[47])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[48])] = value_dict.get('Cable Media 22', None)
	provision_sheet['C' + str(cutnum2[48])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[49])] = value_dict.get('Cable Media 23', None)
	provision_sheet['C' + str(cutnum2[49])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['C' + str(cutnum2[50])] = value_dict.get('Cable Media 24', None)
	provision_sheet['C' + str(cutnum2[50])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	
	#Column D Port Information
	provision_sheet['D' + str(cutnum2[27])] = value_dict.get('Destination Device Port 1', None)
	provision_sheet['D' + str(cutnum2[27])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[28])] = value_dict.get('Destination Device Port 2', None)
	provision_sheet['D' + str(cutnum2[28])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[29])] = value_dict.get('Destination Device Port 3', None)
	provision_sheet['D' + str(cutnum2[29])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[30])] = value_dict.get('Destination Device Port 4', None)
	provision_sheet['D' + str(cutnum2[30])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[31])] = value_dict.get('Destination Device Port 5', None)
	provision_sheet['D' + str(cutnum2[31])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[32])] = value_dict.get('Destination Device Port 6', None)
	provision_sheet['D' + str(cutnum2[32])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[33])] = value_dict.get('Destination Device Port 7', None)
	provision_sheet['D' + str(cutnum2[33])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[34])] = value_dict.get('Destination Device Port 8', None)
	provision_sheet['D' + str(cutnum2[34])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[35])] = value_dict.get('Destination Device Port 9', None)
	provision_sheet['D' + str(cutnum2[35])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[36])] = value_dict.get('Destination Device Port 10', None)
	provision_sheet['D' + str(cutnum2[36])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[37])] = value_dict.get('Destination Device Port 11', None)
	provision_sheet['D' + str(cutnum2[37])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[38])] = value_dict.get('Destination Device Port 12', None)
	provision_sheet['D' + str(cutnum2[38])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[39])] = value_dict.get('Destination Device Port 13', None)
	provision_sheet['D' + str(cutnum2[39])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[40])] = value_dict.get('Destination Device Port 14', None)
	provision_sheet['D' + str(cutnum2[40])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[41])] = value_dict.get('Destination Device Port 15', None)
	provision_sheet['D' + str(cutnum2[41])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[42])] = value_dict.get('Destination Device Port 16', None)
	provision_sheet['D' + str(cutnum2[42])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[43])] = value_dict.get('Destination Device Port 17', None)
	provision_sheet['D' + str(cutnum2[43])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[44])] = value_dict.get('Destination Device Port 18', None)
	provision_sheet['D' + str(cutnum2[44])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[45])] = value_dict.get('Destination Device Port 19', None)
	provision_sheet['D' + str(cutnum2[45])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[46])] = value_dict.get('Destination Device Port 20', None)
	provision_sheet['D' + str(cutnum2[46])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[47])] = value_dict.get('Destination Device Port 21', None)
	provision_sheet['D' + str(cutnum2[47])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[48])] = value_dict.get('Destination Device Port 22', None)
	provision_sheet['D' + str(cutnum2[48])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[49])] = value_dict.get('Destination Device Port 23', None)
	provision_sheet['D' + str(cutnum2[49])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['D' + str(cutnum2[50])] = value_dict.get('Destination Device Port 24', None)
	provision_sheet['D' + str(cutnum2[50])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	
	#Column E Port Information
	
	provision_sheet['E' + str(cutnum2[27])] = value_dict.get('Destination Termination Port 1', None)
	provision_sheet['E' + str(cutnum2[27])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[28])] = value_dict.get('Destination Termination Port 2', None)
	provision_sheet['E' + str(cutnum2[28])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[29])] = value_dict.get('Destination Termination Port 3', None)
	provision_sheet['E' + str(cutnum2[29])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[30])] = value_dict.get('Destination Termination Port 4', None)
	provision_sheet['E' + str(cutnum2[30])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[31])] = value_dict.get('Destination Termination Port 5', None)
	provision_sheet['E' + str(cutnum2[31])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[32])] = value_dict.get('Destination Termination Port 6', None)
	provision_sheet['E' + str(cutnum2[32])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[33])] = value_dict.get('Destination Termination Port 7', None)
	provision_sheet['E' + str(cutnum2[33])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[34])] = value_dict.get('Destination Termination Port 8', None)
	provision_sheet['E' + str(cutnum2[34])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[35])] = value_dict.get('Destination Termination Port 9', None)
	provision_sheet['E' + str(cutnum2[35])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[36])] = value_dict.get('Destination Termination Port 10', None)
	provision_sheet['E' + str(cutnum2[36])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[37])] = value_dict.get('Destination Termination Port 11', None)
	provision_sheet['E' + str(cutnum2[37])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[38])] = value_dict.get('Destination Termination Port 12', None)
	provision_sheet['E' + str(cutnum2[38])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[39])] = value_dict.get('Destination Termination Port 13', None)
	provision_sheet['E' + str(cutnum2[39])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[40])] = value_dict.get('Destination Termination Port 14', None)
	provision_sheet['E' + str(cutnum2[40])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[41])] = value_dict.get('Destination Termination Port 15', None)
	provision_sheet['E' + str(cutnum2[41])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[42])] = value_dict.get('Destination Termination Port 16', None)
	provision_sheet['E' + str(cutnum2[42])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[43])] = value_dict.get('Destination Termination Port 17', None)
	provision_sheet['E' + str(cutnum2[43])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[44])] = value_dict.get('Destination Termination Port 18', None)
	provision_sheet['E' + str(cutnum2[44])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[45])] = value_dict.get('Destination Termination Port 19', None)
	provision_sheet['E' + str(cutnum2[45])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[46])] = value_dict.get('Destination Termination Port 20', None)
	provision_sheet['E' + str(cutnum2[46])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[47])] = value_dict.get('Destination Termination Port 21', None)
	provision_sheet['E' + str(cutnum2[47])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[48])] = value_dict.get('Destination Termination Port 22', None)
	provision_sheet['E' + str(cutnum2[48])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[49])] = value_dict.get('Destination Termination Port 23', None)
	provision_sheet['E' + str(cutnum2[49])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	provision_sheet['E' + str(cutnum2[50])] = value_dict.get('Destination Termination Port 24', None)
	provision_sheet['E' + str(cutnum2[50])].border = Border(left=Side(border_style = 'thin', color = '00000000'))
	
	#Wrap text in the critical comments/notes
	#provision_sheet.cell('A' + str(cutnum2[53])).style.alignment.wrap_text = True
	
	#Increment Index Values for cutnum2
	cutnum2[:] = [i + i_constant for i in cutnum2]
	
	#Clear the dictionary for the next cutsheet
	value_dict = {}
	
#Set Column Widths
provision_sheet.column_dimensions['A'].width = 27.88
provision_sheet.column_dimensions['B'].width = 27.88
provision_sheet.column_dimensions['C'].width = 12.88
provision_sheet.column_dimensions['D'].width = 27.88
provision_sheet.column_dimensions['E'].width = 27.88

#Left Justify the Cut Sheets
for row in provision_sheet:
	for cell in row:
		cell.alignment = Alignment(horizontal = 'left')
		
#fit all columns to one page
provision_sheet.sheet_properties.pageSetUpPr.fitToPage = True
provision_sheet.page_setup.fitToHeight = False

#Confidentiality statement in the footer
provision_sheet.header_footer.center_footer.text = 'The contents of this document are confidential and are for official uses only. All other uses are prohibited. Copyright 2016 by Itascapoint Advisory Services, Inc. All rights reserved.'
provision_sheet.header_footer.center_footer.font_size = 10	
#Save the provision Workbook	
provision_wb.save(provision_name)

#create formatting documentation
print 'Writing Formatting Document'
format_handle = open('Formatting_README.txt', 'a')
format_handle.write('In Excel, you will need to complete the following formatting:\nColumn A, B, D, E will need a width of 27.00, Column c will need a width of 12.\nIf Needed, set the width of a column, right click the column letter and select column width then enter the width value. This only needs to be done if the columns are not already the correct size.\nYou will then need to verify the Fit All Columns to 1 Page in the print parameters\nto make sure that the cut sheets fit to one page.')
format_handle.close()

#Success Notes
print 'Cut Sheets Compiled Successfully!'
ctypes.windll.user32.MessageBoxA(0, 'Cut Sheets Compiled Successfully!', 'Success!', 1)
